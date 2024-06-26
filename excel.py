#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Dec 21 10:01:14 2023

@author: wiktor
"""
##########################################################
# Imports
##########################################################
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font

import os
from calendar import monthrange
import datetime
import time
from freeDays import FREEDAYS

import logging
FORMAT = '%(asctime)s %(levelname)s at line %(lineno)s %(name)s %(funcName)s : %(message)s'
logging.basicConfig(filename='myapp.log', level=logging.DEBUG, format = FORMAT, force=True)
del FORMAT
logger=logging.getLogger(__name__)

##########################################################
# Defines
##########################################################
DIRECTORY = '/home/wiktor/'
ARCHIVE = 'archiwum/'

DATA = "Dane"
HOURS = "Godziny"

TOKEN_ID_COLUMN = 1
SURNAME_COLUMN = 2
RATE_COLUMN = 3
HOURS_COLUMN = 4
ADVANCE_PAYMENT_COLUMN = 5
INSURANCE_COLUMN = 6
SALARY_COLUMN = 7
VACATION_COLUMN = 8
WORKING_STATUS_COLUMN = 9
AMMOUNT_OF_EMPLOYEES_COLUMN = 10
DATE_COLUMN = 12

TITLE_ROW = 1
FIRST_DATA_ROW = 2

WEEKEND_PATTERN = PatternFill(start_color="FFD966",
                                    end_color="FFD966",
                                    fill_type = "solid")

CHRISTMAS_PATTERN = PatternFill(start_color="EB8034",
                                end_color="EB8034",
                                fill_type = "solid")

GREY_PATTERN = PatternFill(start_color="DBDBDB",
                            end_color="DBDBDB",
                            fill_type = "solid")
                            #FFC0C0C0

RED_PATTERN = PatternFill(start_color="FF0000",
                            end_color="FF0000",
                            fill_type = "solid")
                            #

WHITE_PATTERN = PatternFill(start_color='FFFFFF',
                            end_color='FF9933',
                            fill_type = "solid")

ORANGE_PATTERN = PatternFill(start_color='E34829',
                             end_color='E34829',
                             fill_type = "solid")
                            #'FF993300'

BOLDED = Font(name='calibri', bold=True)

RED_FONT = Font(name='calibri', color='FF0000')

BORDER = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

RIGHT_BORDER = Border(left=Side(style='thin'), 
                right=Side(style='medium'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

CENTER = Alignment(horizontal='center')

##########################################################
# Functions
##########################################################
def numberToColumn(number):
    return chr(64 + number)

##########################################################
# Class
##########################################################
class excelSheetC:
    EMPLOYEE_DATA_PATERN = ["Token", "Nazwisko", "Stawka", "Godziny", "Zaliczka",
                        "Ubezpieczenie", "Wypłata", "Urlop", "Pracuje"]
    
    def __init__(self, filename):
        self.filename = filename
        self.file = DIRECTORY + filename
        self.isFileInitializeCorrectly = False
        self.lastModificationTime = ''


    def initializeSheet(self):
        try:
            self.workbook = load_workbook(filename = self.file)
        except:
            logger.error("Inilization Error. Check is file open correctly")
        self.sheet = self.workbook.active


    def saveSheet(self):
        try:
            self.workbook.save(filename = self.file)
        except:
            logger.error("Saving Error")


    def create_workbook(self):
        self.workbook = Workbook()
        self.workbook.create_sheet(DATA)
        self.workbook.create_sheet(HOURS)
        self.workbook.remove(self.workbook['Sheet'])
        self.saveSheet


    def checkFileAndInitialize(self):
        isFileExist = False
        if(os.path.isfile(self.file)):
            self.initializeSheet()
            isFileExist = True
        else:
            logger.error("File dosn't exist: " + str(self.file))
        self.isFileInitializeCorrectly = isFileExist


    def paintRows(self, sheet, rowToEdit, columnToEdit, color, ammountOfColumnsToEdit = 1):
        for i in range (ammountOfColumnsToEdit):
            self.workbook[sheet].cell(row=rowToEdit, column=(columnToEdit + i)).fill = color


    def checkAndPaintGrey(self, sheet, rowToEdit, columnToEdit, ammountOfColumnsToEdit = 1):
        isGreyRow = False
        if ((rowToEdit % 2) == 0):
            self.paintRows(sheet, rowToEdit, columnToEdit, GREY_PATTERN, ammountOfColumnsToEdit)
            isGreyRow = True
        return isGreyRow


    def isWeekendOrChristmas(self, 
                             date,  
                             sheet, 
                             rowToEdit, 
                             columnToEdit, 
                             ammountOfColumnsToEdit = 1):
        if (date.weekday() > 4):
            self.paintRows(sheet, rowToEdit, columnToEdit, WEEKEND_PATTERN, ammountOfColumnsToEdit)
        for i in FREEDAYS:
            christmasToCompare = datetime.datetime.strptime(i, "%Y-%m-%d")
            if (christmasToCompare.date() == date.date()):
                self.paintRows(sheet, rowToEdit, columnToEdit, CHRISTMAS_PATTERN, ammountOfColumnsToEdit)
            elif (christmasToCompare.date() > date.date()):
                break

    def generateHoursTemplate(self, employees, date, daysInMonth):
        # Generating names and token ID in hours sheet
        for i in range(0, employees.ammountOfEmplotees):
            columnToEdit = (i + 1) * 2
            self.workbook[HOURS].cell(row=TITLE_ROW, column=columnToEdit).value = employees.employee[i].name
            self.workbook[HOURS].cell(row=TITLE_ROW, column=columnToEdit).border = BORDER
            self.workbook[HOURS].cell(row=TITLE_ROW, column=columnToEdit+1).value = employees.employee[i].tokenId
            self.workbook[HOURS].cell(row=TITLE_ROW, column=columnToEdit+1).border = RIGHT_BORDER
            self.workbook[HOURS].cell(row=TITLE_ROW, column=columnToEdit+1).number_format = '0'

            # Generating cells to date input
            for j in range(daysInMonth + 1):
                rowToEdit = j + 1
                self.workbook[HOURS].cell(row=rowToEdit, column=columnToEdit).border = BORDER
                self.workbook[HOURS].cell(row=rowToEdit, column=columnToEdit+1).border = RIGHT_BORDER

        # Generating day and month column
        self.workbook[HOURS].cell(row=1, column=1).border = RIGHT_BORDER
        for i in range(daysInMonth):
            rowToEdit = i + 2

            processedDate = datetime.datetime(date.year, date.month, i + 1)
            self.workbook[HOURS].cell(row=rowToEdit, column=1).value = processedDate
                
            self.workbook[HOURS].cell(row=rowToEdit, column=1).number_format = 'DD MMM'
            self.workbook[HOURS].cell(row=rowToEdit, column=1).alignment = CENTER
            self.workbook[HOURS].cell(row=rowToEdit, column=1).border = RIGHT_BORDER

            self.checkAndPaintGrey(HOURS, rowToEdit, 1, (employees.ammountOfEmplotees * 2) + 1)
            self.isWeekendOrChristmas(processedDate,
                                      HOURS, 
                                      rowToEdit, 
                                      1, 
                                      (employees.ammountOfEmplotees * 2) + 1)


    def generateExcelTemplate(self, employees, date):
        
        daysInMonth = monthrange(date.year, date.month)[1]
        self.generateHoursTemplate(employees, date, daysInMonth)
  
        # Ammount of employees cell
        self.workbook[DATA].cell(row=1, column=AMMOUNT_OF_EMPLOYEES_COLUMN).value = "Pracownicy"
        self.workbook[DATA].cell(row=1, column=AMMOUNT_OF_EMPLOYEES_COLUMN).border = BORDER

        self.workbook[DATA].cell(row=2, column=AMMOUNT_OF_EMPLOYEES_COLUMN).value = \
            employees.ammountOfEmplotees
        self.workbook[DATA].cell(row=2, column=AMMOUNT_OF_EMPLOYEES_COLUMN).border = BORDER
        self.checkAndPaintGrey(DATA, 2, AMMOUNT_OF_EMPLOYEES_COLUMN) 

        # Generating hours section
        # Generating 'DATA' cell
        self.workbook[DATA].cell(row=1, column=DATE_COLUMN).value = "Data"
        self.workbook[DATA].cell(row=1, column=DATE_COLUMN).font = BOLDED
        self.workbook[DATA].cell(row=1, column=DATE_COLUMN).border = BORDER
        self.workbook[DATA].cell(row=1, column=DATE_COLUMN).alignment = CENTER

        # Generating row
        for i in range(employees.ammountOfEmplotees):
            columnToEdit = DATE_COLUMN + (i + 1)

            self.workbook[DATA].cell(row=1, column=columnToEdit).value = \
                employees.employee[i].name
            self.workbook[DATA].cell(row=1, column=columnToEdit).font = BOLDED
            self.workbook[DATA].cell(row=1, column=columnToEdit).border = BORDER

        # Generating dates and month column
        for i in range(daysInMonth):
            rowToEdit = i + 2

            processedDate = datetime.datetime(date.year, date.month, i + 1)
            self.workbook[DATA].cell(row=rowToEdit, column=DATE_COLUMN).value = processedDate
                
            self.workbook[DATA].cell(row=rowToEdit, column=DATE_COLUMN).number_format = 'DD MMM'
            self.workbook[DATA].cell(row=rowToEdit, column=DATE_COLUMN).border = BORDER
            self.workbook[DATA].cell(row=rowToEdit, column=DATE_COLUMN).alignment = CENTER

            
            self.checkAndPaintGrey(DATA, rowToEdit, DATE_COLUMN, employees.ammountOfEmplotees + 1)
            self.isWeekendOrChristmas(processedDate,
                                      DATA, 
                                      rowToEdit, 
                                      DATE_COLUMN + 1, 
                                      employees.ammountOfEmplotees)


            for j in range(employees.ammountOfEmplotees):
                startColumnInHoursSheet = numberToColumn((j + 1) * 2)
                endColumnInHoursSheet = numberToColumn(((j + 1) * 2) + 1)

                columnToEdit = DATE_COLUMN + (j + 1)
                valueToInput = "=" + HOURS + "!" + endColumnInHoursSheet + str(rowToEdit) + \
                    "-" + HOURS + "!" + startColumnInHoursSheet + str(rowToEdit)
                    
                self.workbook[DATA].cell(row=rowToEdit, column=columnToEdit).value = valueToInput
                self.workbook[DATA].cell(row=rowToEdit, column=columnToEdit).border = BORDER
                self.workbook[DATA].cell(row=rowToEdit, column=columnToEdit).number_format = 'h:mm'

        # Adding employees data
        for i in range(len(self.EMPLOYEE_DATA_PATERN)):
            self.workbook[DATA].cell(row=1, column=i+1).value = self.EMPLOYEE_DATA_PATERN[i]
            self.workbook[DATA].cell(row=1, column=i+1).border = BORDER

        for i in range(employees.ammountOfEmplotees):
            editedRow = i + 2

            self.workbook[DATA].cell(row=editedRow, column=TOKEN_ID_COLUMN).number_format = '0'
            self.workbook[DATA].cell(row=editedRow, column=TOKEN_ID_COLUMN).value = employees.employee[i].tokenId
            self.workbook[DATA].cell(row=editedRow, column=TOKEN_ID_COLUMN).border = BORDER

            self.workbook[DATA].cell(row=editedRow, column=SURNAME_COLUMN).value = employees.employee[i].name
            self.workbook[DATA].cell(row=editedRow, column=SURNAME_COLUMN).border = BORDER

            self.workbook[DATA].cell(row=editedRow, column=RATE_COLUMN).value = employees.employee[i].rate
            self.workbook[DATA].cell(row=editedRow, column=RATE_COLUMN).border = BORDER

            employeeWorkTimeColumn = numberToColumn(DATE_COLUMN + i + 1)
            valueToInput = "=SUM(" + employeeWorkTimeColumn + "2:" + employeeWorkTimeColumn + \
                  str(daysInMonth + 1) + ")"
            self.workbook[DATA].cell(row=editedRow, column=HOURS_COLUMN).value = valueToInput
            self.workbook[DATA].cell(row=editedRow, column=HOURS_COLUMN).border = BORDER
            self.workbook[DATA].cell(row=editedRow, column=HOURS_COLUMN).number_format = '[h]:mm'

            self.workbook[DATA].cell(row=editedRow, column=ADVANCE_PAYMENT_COLUMN).border = BORDER
            self.workbook[DATA].cell(row=editedRow, column=ADVANCE_PAYMENT_COLUMN).number_format = \
                '#,##0.00"zł"' 
            
            self.workbook[DATA].cell(row=editedRow, column=INSURANCE_COLUMN).border = BORDER
            self.workbook[DATA].cell(row=editedRow, column=INSURANCE_COLUMN).number_format = \
                '#,##0.00"zł"'
            
            self.workbook[DATA].cell(row=editedRow, column=SALARY_COLUMN).border = BORDER
            self.workbook[DATA].cell(row=editedRow, column=SALARY_COLUMN).number_format = \
                '#,##0.00"zł"'
            self.workbook[DATA].cell(row=editedRow, column=SALARY_COLUMN).font = RED_FONT
            strI = str(i + 2)
            valueToInput = "=((" + numberToColumn(HOURS_COLUMN) + strI + "*24)*" + \
                numberToColumn(RATE_COLUMN) + strI + ")-" + \
                numberToColumn(ADVANCE_PAYMENT_COLUMN) + strI + "-" + \
                numberToColumn(INSURANCE_COLUMN) + strI
            self.workbook[DATA].cell(row=editedRow, column=SALARY_COLUMN).value = valueToInput
            
            self.workbook[DATA].cell(row=editedRow, column=VACATION_COLUMN).border = BORDER                

            self.workbook[DATA].cell(row=editedRow, column=WORKING_STATUS_COLUMN).value = "T"
            self.workbook[DATA].cell(row=editedRow, column=WORKING_STATUS_COLUMN).border = BORDER

            self.checkAndPaintGrey(DATA, editedRow, TOKEN_ID_COLUMN, WORKING_STATUS_COLUMN)        


    def addNewEmployeeCells(self, employeeNumber, employee, date):
        rowToEdit = employeeNumber + 2
        columnToEditInHoursSheet = ((employeeNumber + 1) * 2)
        daysInMonth = monthrange(date.year, date.month)[1]

        startColumnInHoursSheet = numberToColumn(columnToEditInHoursSheet)
        endColumnInHoursSheet = numberToColumn(columnToEditInHoursSheet + 1) # TO DO change to start + 1

        self.workbook[HOURS].cell(row=TITLE_ROW, column=columnToEditInHoursSheet).value = employee.name
        self.workbook[HOURS].cell(row=TITLE_ROW, column=columnToEditInHoursSheet).border = BORDER
        self.workbook[HOURS].cell(row=TITLE_ROW, column=(columnToEditInHoursSheet + 1)).value = employee.tokenId
        self.workbook[HOURS].cell(row=TITLE_ROW, column=(columnToEditInHoursSheet + 1)).border = RIGHT_BORDER
        self.workbook[HOURS].cell(row=TITLE_ROW, column=(columnToEditInHoursSheet + 1)).number_format = '0'

        self.workbook[DATA].cell(row=rowToEdit, column=TOKEN_ID_COLUMN).number_format = '0'                
        
        self.checkAndPaintGrey(DATA, rowToEdit, TOKEN_ID_COLUMN, WORKING_STATUS_COLUMN)
        employeeWorkTimeColumn = numberToColumn(DATE_COLUMN + employeeNumber + 1)
        valueToInput = "=SUM(" + employeeWorkTimeColumn + "2:" + employeeWorkTimeColumn + \
                str(daysInMonth + 1) + ")"
        self.workbook[DATA].cell(row=rowToEdit, column=HOURS_COLUMN).value = valueToInput
        self.workbook[DATA].cell(row=rowToEdit, column=HOURS_COLUMN).border = BORDER
        self.workbook[DATA].cell(row=rowToEdit, column=HOURS_COLUMN).number_format = '[h]:mm'

        self.workbook[DATA].cell(row=rowToEdit, column=ADVANCE_PAYMENT_COLUMN).border = BORDER
        self.workbook[DATA].cell(row=rowToEdit, column=ADVANCE_PAYMENT_COLUMN).number_format = \
            '#,##0.00"zł"' 
        
        self.workbook[DATA].cell(row=rowToEdit, column=INSURANCE_COLUMN).border = BORDER
        self.workbook[DATA].cell(row=rowToEdit, column=INSURANCE_COLUMN).number_format = \
            '#,##0.00"zł"'
        
        self.workbook[DATA].cell(row=rowToEdit, column=SALARY_COLUMN).border = BORDER
        self.workbook[DATA].cell(row=rowToEdit, column=SALARY_COLUMN).number_format = \
            '#,##0.00"zł"'
        self.workbook[DATA].cell(row=rowToEdit, column=SALARY_COLUMN).font = RED_FONT
        rowToEditInStr = str(rowToEdit)
        valueToInput = "=((" + numberToColumn(HOURS_COLUMN) + rowToEditInStr + "*24)*" + \
            numberToColumn(RATE_COLUMN) + rowToEditInStr + ")-" + \
            numberToColumn(ADVANCE_PAYMENT_COLUMN) + rowToEditInStr + "-" + \
            numberToColumn(INSURANCE_COLUMN) + rowToEditInStr
        self.workbook[DATA].cell(row=rowToEdit, column=SALARY_COLUMN).value = valueToInput
        
        self.workbook[DATA].cell(row=rowToEdit, column=VACATION_COLUMN).border = BORDER                

        self.workbook[DATA].cell(row=rowToEdit, column=WORKING_STATUS_COLUMN).value = "T"
        self.workbook[DATA].cell(row=rowToEdit, column=WORKING_STATUS_COLUMN).border = BORDER

        for i in range(TOKEN_ID_COLUMN, RATE_COLUMN + 1):
            self.workbook[DATA].cell(row=rowToEdit, column=i).border = BORDER
        
        columnToEdit = DATE_COLUMN + (employeeNumber + 1)
        self.workbook[DATA].cell(row=TITLE_ROW, column=columnToEdit).value = \
            employee.name
        self.workbook[DATA].cell(row=TITLE_ROW, column=columnToEdit).font = BOLDED
        self.workbook[DATA].cell(row=TITLE_ROW, column=columnToEdit).border = BORDER
          

        for i in range(daysInMonth):
            rowToEdit = i + 2
            processedDate = datetime.datetime(date.year, date.month, i + 1)

            self.workbook[HOURS].cell(row=rowToEdit, column=columnToEditInHoursSheet).border = BORDER
            self.workbook[HOURS].cell(row=rowToEdit, column=(columnToEditInHoursSheet + 1)).border = RIGHT_BORDER
            self.checkAndPaintGrey(HOURS, rowToEdit, columnToEditInHoursSheet, 2)
            self.isWeekendOrChristmas(processedDate,
                                      HOURS, 
                                      rowToEdit, 
                                      columnToEditInHoursSheet,
                                      2)

            valueToInput = "=" + HOURS + "!" + endColumnInHoursSheet + str(rowToEdit) + \
                "-" + HOURS + "!" + startColumnInHoursSheet + str(rowToEdit)
                
            self.workbook[DATA].cell(row=rowToEdit, column=columnToEdit).value = valueToInput
            self.workbook[DATA].cell(row=rowToEdit, column=columnToEdit).border = BORDER
            self.workbook[DATA].cell(row=rowToEdit, column=columnToEdit).number_format = 'h:mm'

            self.checkAndPaintGrey(DATA, rowToEdit, columnToEdit)
            self.isWeekendOrChristmas(processedDate,
                                      DATA, 
                                      rowToEdit, 
                                      columnToEdit)


    def updateAmmountOfEmployees(self, ammountOfEmplotees):
        tmp = self.workbook[DATA].cell(row=2, column=AMMOUNT_OF_EMPLOYEES_COLUMN).value
        self.workbook[DATA].cell(row=2, column=AMMOUNT_OF_EMPLOYEES_COLUMN).value = tmp + 1    


    def inputTimestampIntoExcel(self, id, employeeId, currentTime):
        employeeColumn = (2 * (employeeId + 1))
        if (self.workbook[HOURS].cell(row=TITLE_ROW, column=(employeeColumn + 1)).value == id):
            dateRow = currentTime.day + TITLE_ROW
            columnDate = self.workbook[HOURS].cell(row=dateRow, column=1).value.strftime("%d:%m")
            if (columnDate == currentTime.strftime("%d:%m")):
                tmp = self.workbook[HOURS].cell(row=dateRow, column=employeeColumn).value
                
                if (tmp == None):
                    self.workbook[HOURS].cell(row=dateRow, column=employeeColumn).value = currentTime
                    self.workbook[HOURS].cell(row=dateRow, column=employeeColumn).number_format = 'h:mm'

                    self.workbook[HOURS].cell(row=dateRow, column=(employeeColumn + 1)).value = currentTime
                    self.workbook[HOURS].cell(row=dateRow, column=(employeeColumn + 1)).number_format = 'h:mm'
                else:
                    self.workbook[HOURS].cell(row=dateRow, column=(employeeColumn + 1)).value = currentTime

                self.saveSheet()
                self.lastModificationTime = time.ctime(os.path.getmtime(self.file))
            else:
                logger.error("Invalid date in 'HOURS' sheet")
        else:
            logger.error("Invalid TokenId in 'HOURS' sheet")
            
        
    def inputTokenIdToExce(self, id):
        self.workbook[DATA].cell(row=4, column=AMMOUNT_OF_EMPLOYEES_COLUMN).value = id
        self.workbook[DATA].cell(row=4, column=AMMOUNT_OF_EMPLOYEES_COLUMN).number_format = '0'
        self.saveSheet()
        self.lastModificationTime = time.ctime(os.path.getmtime(self.file))
